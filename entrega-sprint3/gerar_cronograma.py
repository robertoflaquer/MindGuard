"""
Gera Q2-Cronograma.xlsx para a entrega Sprint 3 - FIAP PCP
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Cronograma MindGuard"

# ── Paleta de cores ────────────────────────────────────────────────
AZUL_ESCURO   = "1565C0"
AZUL_MEDIO    = "1976D2"
AZUL_CLARO    = "BBDEFB"
VERDE_ESCURO  = "2E7D32"
VERDE_CLARO   = "C8E6C9"
AMARELO       = "FFF9C4"
AMARELO_BORDA = "F9A825"
CINZA_HEADER  = "ECEFF1"
CINZA_LINHA   = "F5F5F5"
BRANCO        = "FFFFFF"
LARANJA       = "FF6F00"
LARANJA_CLARO = "FFE0B2"
VERMELHO      = "B71C1C"
VERMELHO_CLARO= "FFCDD2"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def center():
    return Alignment(horizontal="center", vertical="center",
                     wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center",
                     wrap_text=True)

def border_thin():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium():
    s = Side(style="medium", color="9E9E9E")
    return Border(left=s, right=s, top=s, bottom=s)

# ── Título principal ──────────────────────────────────────────────
ws.merge_cells("A1:L1")
ws["A1"] = "MINDGUARD — Cronograma de Desenvolvimento"
ws["A1"].font = font(bold=True, color="FFFFFF", size=14)
ws["A1"].fill = fill(AZUL_ESCURO)
ws["A1"].alignment = center()
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:L2")
ws["A2"] = ("FIAP · 1º Ano Ciência da Computação · PCP – Prof. Allan Roberto Molto · "
            "2º Semestre Sprint 3 · Roberto Flaquer Moura")
ws["A2"].font = font(italic=True, color="FFFFFF", size=10)
ws["A2"].fill = fill(AZUL_MEDIO)
ws["A2"].alignment = center()
ws.row_dimensions[2].height = 18

# ── Linha em branco ───────────────────────────────────────────────
ws.row_dimensions[3].height = 8

# ── Cabeçalho da tabela ───────────────────────────────────────────
HEADERS = [
    "Fase", "Etapa / Atividade", "Período", "Duração",
    "Responsável", "Status",
    "Fev\n2026", "Mar\n2026", "Abr\n2026",
    "Mai\n2026", "Jun\n2026", "Jul–Ago\n2026"
]
COLS = list("ABCDEFGHIJKL")

for col_idx, header in enumerate(HEADERS, 1):
    cell = ws.cell(row=4, column=col_idx, value=header)
    cell.font = font(bold=True, color="FFFFFF", size=9)
    cell.fill = fill(AZUL_ESCURO)
    cell.alignment = center()
    cell.border = border_medium()
ws.row_dimensions[4].height = 30

# ── Larguras das colunas ─────────────────────────────────────────
col_widths = [6, 44, 18, 10, 14, 14, 5, 5, 5, 5, 5, 5]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── Dados do cronograma ───────────────────────────────────────────
# (fase, atividade, periodo, duracao, responsavel, status,
#  fev, mar, abr, mai, jun, jul-ago)
# Status: CONCLUÍDO | EM ANDAMENTO | PLANEJADO
# Células de período (G-L): "█" se ativa naquele mês, "" se não

ROWS = [
    # FASE 0
    ("0", "PESQUISA & PLANEJAMENTO", "", "", "", "",
     "", "", "", "", "", ""),

    ("0.1", "Análise do kick-off Care Plus / Blua e definição do escopo",
     "Fev 2026", "1 sem", "Roberto", "CONCLUÍDO",
     "█", "", "", "", "", ""),

    ("0.2", "Escolha da stack tecnológica (React, Node, Python, PostgreSQL)",
     "Fev 2026", "1 sem", "Roberto", "CONCLUÍDO",
     "█", "", "", "", "", ""),

    ("0.3", "Design da arquitetura de 4 camadas (Frontend / Backend / Python / DB)",
     "Fev 2026", "1 sem", "Roberto", "CONCLUÍDO",
     "█", "", "", "", "", ""),

    # FASE 1
    ("1", "INFRAESTRUTURA", "", "", "", "",
     "", "", "", "", "", ""),

    ("1.1", "Schema PostgreSQL completo (28+ tabelas: users, signals, risk, specialists, appointments, prescriptions…)",
     "Fev–Mar 2026", "2 sem", "Roberto", "CONCLUÍDO",
     "█", "█", "", "", "", ""),

    ("1.2", "Seed data (signal_types, questionnaire_types, risk_levels, actions, specialists)",
     "Mar 2026", "3 dias", "Roberto", "CONCLUÍDO",
     "", "█", "", "", "", ""),

    ("1.3", "Docker + docker-compose para todos os serviços",
     "Mar 2026", "1 sem", "Roberto", "CONCLUÍDO",
     "", "█", "", "", "", ""),

    # FASE 2
    ("2", "BACKEND NODE.JS (API REST)", "", "", "", "",
     "", "", "", "", "", ""),

    ("2.1", "Auth: POST /register, POST /login, JWT 7 dias, bcrypt, rate limiting",
     "Mar 2026", "1 sem", "Roberto", "CONCLUÍDO",
     "", "█", "", "", "", ""),

    ("2.2", "Signals API: POST /batch, GET /recent, GET /types, GET /stats, POST /simulate",
     "Mar 2026", "1 sem", "Roberto", "CONCLUÍDO",
     "", "█", "", "", "", ""),

    ("2.3", "Risk API: GET /current, GET /history, POST /assess",
     "Mar–Abr 2026", "1 sem", "Roberto", "CONCLUÍDO",
     "", "█", "█", "", "", ""),

    ("2.4", "Questionnaires API: GET /due, POST /submit (score PSS/CBI/OLBI/GAD-7 + reversão), GET /history",
     "Abr 2026", "1 sem", "Roberto", "CONCLUÍDO",
     "", "", "█", "", "", ""),

    ("2.5", "Contexts API: GET /types, GET /active, POST, PATCH /:id/close",
     "Abr 2026", "1 sem", "Roberto", "CONCLUÍDO",
     "", "", "█", "", "", ""),

    ("2.6", "Segurança: XSS sanitization (xss-clean), SQL injection fix (INTERVAL parameterizado), logs estruturados Pino",
     "Mai 2026", "2 dias", "Roberto", "CONCLUÍDO",
     "", "", "", "█", "", ""),

    # FASE 3
    ("3", "PYTHON ENGINE (FastAPI — IA de Análise)", "", "", "", "",
     "", "", "", "", "", ""),

    ("3.1", "BaselineCalculator: mediana + IQR, rolling 7d e 14d, upsert sem quebrar FKs",
     "Abr 2026", "1 sem", "Roberto", "CONCLUÍDO",
     "", "", "█", "", "", ""),

    ("3.2", "DeviationAnalyzer: % desvio vs baseline, threshold 15%, direção",
     "Abr 2026", "1 sem", "Roberto", "CONCLUÍDO",
     "", "", "█", "", "", ""),

    ("3.3", "CorrelationEngine: convergência de sinais, weighted score",
     "Abr 2026", "1 sem", "Roberto", "CONCLUÍDO",
     "", "", "█", "", "", ""),

    ("3.4", "RiskScorer: PSS (38%) + CBI (22%) + OLBI (17%) + GAD-7 (13%) + Check-in (10%) + sinais (30%)",
     "Abr 2026", "1 sem", "Roberto", "CONCLUÍDO",
     "", "", "█", "", "", ""),

    ("3.5", "Fallback por thresholds populacionais (sem baseline ainda)",
     "Abr 2026", "2 dias", "Roberto", "CONCLUÍDO",
     "", "", "█", "", "", ""),

    ("3.6", "Correção de bugs: secondary_factors TEXT[], convergence_count, sinais bons",
     "Abr–Mai 2026", "1 sem", "Roberto", "CONCLUÍDO",
     "", "", "█", "█", "", ""),

    # FASE 4
    ("4", "FRONTEND WEB (React + Vite)", "", "", "", "",
     "", "", "", "", "", ""),

    ("4.1", "Login e Register com validação + JWT persistido em localStorage",
     "Abr 2026", "1 sem", "Roberto", "CONCLUÍDO",
     "", "", "█", "", "", ""),

    ("4.2", "Dashboard: tabs Visão Geral / Registrar Sinais, RiskCard, SignalChart, botão Simular Wearable",
     "Abr–Mai 2026", "2 sem", "Roberto", "CONCLUÍDO",
     "", "", "█", "█", "", ""),

    ("4.3", "Formulários de questionários: PSS-10, CBI, OLBI, Daily Check-in",
     "Abr–Mai 2026", "2 sem", "Roberto", "CONCLUÍDO",
     "", "", "█", "█", "", ""),

    ("4.4", "Páginas: Contextos de Vida (10 tipos + personalizado) + Prescrições (accordion + hash SHA-256)",
     "Mai 2026", "1 sem", "Roberto", "CONCLUÍDO",
     "", "", "", "█", "", ""),

    ("4.5", "Formulário GAD-7: 7 perguntas, score 0–21, 5 níveis de ansiedade, preview em tempo real",
     "Mai 2026", "2 dias", "Roberto", "CONCLUÍDO",
     "", "", "", "█", "", ""),

    # FASE 5
    ("5", "DESIGN SYSTEM & UX (CarePlus / Blua inspired)", "", "", "", "",
     "", "", "", "", "", ""),

    ("5.1", "CSS Variables dark/light mode + useThemeStore (sem flash no carregamento)",
     "Mai 2026", "3 dias", "Roberto", "CONCLUÍDO",
     "", "", "", "█", "", ""),

    ("5.2", "Fonte Nunito, paleta azul royal, componentes .card .btn-primary .input-field",
     "Mai 2026", "3 dias", "Roberto", "CONCLUÍDO",
     "", "", "", "█", "", ""),

    ("5.3", "RiskCard com gauge circular SVG + explicação em português",
     "Mai 2026", "2 dias", "Roberto", "CONCLUÍDO",
     "", "", "", "█", "", ""),

    ("5.4", "Tradução dos sinais para português, correção header modo claro",
     "Mai 2026", "1 dia", "Roberto", "CONCLUÍDO",
     "", "", "", "█", "", ""),

    # FASE 6
    ("6", "AGENDAMENTO DE TELECONSULTA (Backend + Frontend)", "", "", "", "",
     "", "", "", "", "", ""),

    ("6.1", "Tabelas specialists + appointments no PostgreSQL com seed de 3 especialistas",
     "Mai 2026", "2 dias", "Roberto", "CONCLUÍDO",
     "", "", "", "█", "", ""),

    ("6.2", "Appointments API: GET /specialists, GET /slots, POST /, GET /, PATCH /:id/cancel",
     "Mai 2026", "3 dias", "Roberto", "CONCLUÍDO",
     "", "", "", "█", "", ""),

    ("6.3", "Página /treatment: especialista + calendário 7 dias + grade de horários + confirmação",
     "Mai 2026", "3 dias", "Roberto", "CONCLUÍDO",
     "", "", "", "█", "", ""),

    ("6.4", "useAppointmentStore (Zustand) + Treatment.jsx integrado com API real + conflito de horário",
     "Mai 2026", "2 dias", "Roberto", "CONCLUÍDO",
     "", "", "", "█", "", ""),

    # FASE 7
    ("7", "PESQUISA DE WEARABLES", "", "", "", "",
     "", "", "", "", "", ""),

    ("7.1", "Pesquisa HealthKit (Apple Watch): HKObserverQuery + Background Delivery",
     "Mai 2026", "1 dia", "Roberto", "CONCLUÍDO",
     "", "", "", "█", "", ""),

    ("7.2", "Pesquisa Android Health Connect + Samsung Health Data SDK",
     "Mai 2026", "1 dia", "Roberto", "CONCLUÍDO",
     "", "", "", "█", "", ""),

    ("7.3", "Avaliação de plataformas: Open Wearables (gratuito) vs Terra API ($399/mês)",
     "Mai 2026", "1 dia", "Roberto", "CONCLUÍDO",
     "", "", "", "█", "", ""),

    ("7.4", "Endpoint POST /api/signals/simulate (7 sinais realistas, Apple Watch e Galaxy Watch)",
     "Mai 2026", "1 dia", "Roberto", "CONCLUÍDO",
     "", "", "", "█", "", ""),

    # FASE 8
    ("8", "PRESCRIÇÃO DIGITAL PÓS-TELECONSULTA", "", "", "", "",
     "", "", "", "", "", ""),

    ("8.1", "Tabela prescriptions com hash SHA-256 + Prescriptions API: POST /, GET /, GET /:id",
     "Mai 2026", "2 dias", "Roberto", "CONCLUÍDO",
     "", "", "", "█", "", ""),

    ("8.2", "Página /prescriptions: lista com accordion, itens, observações, hash de auditoria visual",
     "Mai 2026", "2 dias", "Roberto", "CONCLUÍDO",
     "", "", "", "█", "", ""),

    ("8.3", "Geração de PDF com assinatura digital (PDFKit)",
     "Jun 2026", "1 sem", "Roberto", "PLANEJADO",
     "", "", "", "", "█", ""),

    ("8.4", "Role médico no sistema de auth + interface médica de prescrição",
     "Jun 2026", "1 sem", "Roberto", "PLANEJADO",
     "", "", "", "", "█", ""),

    # FASE 9
    ("9", "LIMPEZA DE CÓDIGO", "", "", "", "",
     "", "", "", "", "", ""),

    ("9.1", "print() no Python Engine → logging estruturado; console.log no backend → Pino",
     "Mai 2026", "1 dia", "Roberto", "CONCLUÍDO",
     "", "", "", "█", "", ""),

    ("9.2", "Código morto removido: optionalAuth, getTotalSignalCount, fetchRiskLevels, getStats",
     "Mai 2026", "1 dia", "Roberto", "CONCLUÍDO",
     "", "", "", "█", "", ""),

    ("9.3", "README.md criado nos 3 projetos (backend, frontend, python-engine)",
     "Mai 2026", "1 dia", "Roberto", "CONCLUÍDO",
     "", "", "", "█", "", ""),

    # FASE 10 — PLANEJADO
    ("10", "DEPLOY EM PRODUÇÃO", "", "", "", "",
     "", "", "", "", "", ""),

    ("10.1", "Testar docker-compose up --build localmente (todos 4 serviços)",
     "Jun 2026", "2 dias", "Roberto", "PLANEJADO",
     "", "", "", "", "█", ""),

    ("10.2", "Configurar .env de produção + deploy Railway.app ou Render.com com HTTPS",
     "Jun 2026", "3 dias", "Roberto", "PLANEJADO",
     "", "", "", "", "█", ""),

    ("10.3", "Testes end-to-end em produção + configuração de domínio",
     "Jun 2026", "2 dias", "Roberto", "PLANEJADO",
     "", "", "", "", "█", ""),

    # FASE 11 — PLANEJADO
    ("11", "APP MOBILE + WEARABLES FÍSICOS", "", "", "", "",
     "", "", "", "", "", ""),

    ("11.1", "App iOS (Swift/SwiftUI): HealthKit + HKObserverQuery + Background Delivery matinal",
     "Jul 2026", "3 sem", "Roberto", "PLANEJADO",
     "", "", "", "", "", "█"),

    ("11.2", "App Android (Kotlin + Jetpack Compose): Health Connect SDK + WorkManager 7h",
     "Jul–Ago 2026", "3 sem", "Roberto", "PLANEJADO",
     "", "", "", "", "", "█"),
]

# ── Cores por status ─────────────────────────────────────────────
STATUS_STYLE = {
    "CONCLUÍDO":     (VERDE_ESCURO,  VERDE_CLARO),
    "EM ANDAMENTO":  (LARANJA,       LARANJA_CLARO),
    "PLANEJADO":     (AZUL_ESCURO,   AZUL_CLARO),
    "":              (AZUL_ESCURO,   CINZA_HEADER),   # cabeçalho de fase
}

# ── Cor das células de "barra" Gantt ─────────────────────────────
GANTT_STATUS = {
    "CONCLUÍDO":    "43A047",   # verde médio
    "EM ANDAMENTO": "FB8C00",   # laranja
    "PLANEJADO":    "1E88E5",   # azul
    "":             "",
}

current_row = 5
for row_data in ROWS:
    fase, ativ, periodo, dur, resp, status = row_data[:6]
    gantt = row_data[6:]  # 6 colunas de meses

    is_header = (ativ == "" or (ativ.isupper() and dur == ""))

    # Detecta linha de fase (sem atividade preenchida ou tudo maiúsculo)
    is_fase_header = (fase.isdigit() or (len(fase) <= 2 and "." not in fase))
    is_fase_header = is_fase_header and dur == ""

    row_height = 28 if is_fase_header else 22
    ws.row_dimensions[current_row].height = row_height

    # ── Células A–F (dados textuais) ──────────────────────────────
    values = [fase, ativ, periodo, dur, resp, status]
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=val)
        cell.border = border_thin()

        if is_fase_header:
            # Linha de cabeçalho de fase
            cell.fill = fill(AZUL_CLARO)
            cell.font = font(bold=True, color=AZUL_ESCURO, size=9)
            cell.alignment = left() if col_idx == 2 else center()
        else:
            # Linha normal — alterna cinza/branco
            bg = CINZA_LINHA if (current_row % 2 == 0) else BRANCO
            cell.fill = fill(bg)
            # Status colorido
            if col_idx == 6 and status:
                font_color, bg_color = STATUS_STYLE.get(status, ("000000", BRANCO))
                cell.fill = fill(bg_color)
                cell.font = font(bold=True, color=font_color, size=8)
            else:
                cell.font = font(size=9) if col_idx != 1 else font(bold=True, size=9, color=AZUL_ESCURO)
            cell.alignment = left() if col_idx == 2 else center()

    # ── Células G–L (Gantt) ───────────────────────────────────────
    gantt_color = GANTT_STATUS.get(status, "")
    for col_idx, val in enumerate(gantt, 7):
        cell = ws.cell(row=current_row, column=col_idx, value=val)
        cell.border = border_thin()
        cell.alignment = center()
        if val == "█" and gantt_color:
            cell.fill = fill(gantt_color)
            cell.font = font(color=gantt_color, size=9)  # invisível — é o fill que conta
        elif is_fase_header:
            cell.fill = fill(AZUL_CLARO)
        else:
            bg = CINZA_LINHA if (current_row % 2 == 0) else BRANCO
            cell.fill = fill(bg)

    current_row += 1

# ── Legenda ───────────────────────────────────────────────────────
ws.row_dimensions[current_row].height = 8
current_row += 1

legends = [
    ("CONCLUÍDO", VERDE_ESCURO, VERDE_CLARO),
    ("EM ANDAMENTO", LARANJA, LARANJA_CLARO),
    ("PLANEJADO", AZUL_ESCURO, AZUL_CLARO),
]
ws.merge_cells(f"A{current_row}:B{current_row}")
ws[f"A{current_row}"] = "LEGENDA:"
ws[f"A{current_row}"].font = font(bold=True, size=9)
ws[f"A{current_row}"].alignment = left()

col_offset = 3
for label, fc, bc in legends:
    cell = ws.cell(row=current_row, column=col_offset, value=f"  {label}  ")
    cell.font = font(bold=True, color=fc, size=8)
    cell.fill = fill(bc)
    cell.alignment = center()
    cell.border = border_thin()
    col_offset += 1

ws.row_dimensions[current_row].height = 18
current_row += 1

# Linha de barras Gantt coloridas (legenda visual)
ws.merge_cells(f"A{current_row}:B{current_row}")
ws[f"A{current_row}"] = "Barras Gantt:"
ws[f"A{current_row}"].font = font(bold=True, size=9)
for col_idx, (_, fc, bc) in enumerate(legends, 3):
    cell = ws.cell(row=current_row, column=col_idx)
    cell.fill = fill(fc)
    cell.border = border_thin()
ws.row_dimensions[current_row].height = 14

# ── Freeze panes + filtros ────────────────────────────────────────
ws.freeze_panes = "B5"
ws.auto_filter.ref = f"A4:{get_column_letter(len(HEADERS))}4"

# ── Largura final da coluna A ─────────────────────────────────────
ws.column_dimensions["A"].width = 6

# ── Salvar ────────────────────────────────────────────────────────
output_path = r"C:\Users\betof\mindguard\entrega-sprint3\Q2-Cronograma.xlsx"
wb.save(output_path)
print(f"Salvo em: {output_path}")
